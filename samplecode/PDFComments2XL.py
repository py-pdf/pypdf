#!/usr/bin/python3
"""
   test/demo program tha extract comments from an pdf into a Excel
   command line:
   PDFComments2XL [-d] [-o output.xls] [input.pdf]
   -d: open Excel output at the end of extraction
   -o: prode the output Excel name/path ; if not present the file is created
       in temp folder named "comments on **PDFfile**.xlsx"
    if no parameters (mainly for idle test), the pdf filename is asked for
"""
from collections import OrderedDict
from datetime import datetime
import sys
import os
import pypdf as PDF;
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

import locale
locale.setlocale(locale.LC_ALL,locale.getdefaultlocale()[0])
    
def ListOutlines(pdfS,outl=None):
    """
    provide as a list of the outlines as tuple Title,Page(0 based),Vertical position in %
    """
    if outl is None:
        lst=[('-',0,0),]
        outl=pdfS.getOutlines()
    else:
        lst=[]
    if isinstance(outl,list):
        for k in outl:
            lst+=ListOutlines(pdfS,k)
    else:
        try:
            top=outl['/Top']
        except:
            top=0
        try:
            pp=pdfS.MyPages[outl.page.idnum]
            lst.append((outl.title,pp[0],100.0*(1.0-float(top/pp[1]))))
        except:
            print("trouble with page idnum",outl.page.idnum)
    return lst

def ListAnnots(pdfS):
    """
    provide as a list of the comments with the response saved in .irt_str field, the list is indexed with idnums
    """
    lst=OrderedDict()
    for pn in range(pdfS.numPages):
        p=pdfS.getPage(pn)
        try:
            a=p.get('/Annots' ).getObject()
            if not isinstance(a,list): a=[a]
            for b in a:
                o=b.getObject()
                if o['/Subtype']=='/Text':
                    try: o['/P']        # le champs '/P' etant optionnel on le reconstruit...
                    except:
                        o.update({PDF.NameObject('/P'):p.indirectRef})
                    o.irt={}
                    lst[b.idnum]=o
        except:
            pass
    #copy the information into the original comment
    for k,o in lst.items():
        if '/IRT' in o:
            t=o['/Contents']
            if isinstance(t,bytes):t=t.replace(b'\r',b'\n').decode('unicode_escape')
            lst[o.rawGet('/IRT').idnum].irt[o['/M']]=\
                            '%s (%s):\n%s'%\
                                (o['/T'],datetime.strptime(o['/M'][2:10],'%Y%m%d').strftime('%x'),t)
    #concat all replied comments into one string to ease insertion later...
    for o in lst.values():
        o.irt_str='\n'.join([o.irt[x] for x in sorted(o.irt.keys())])
    return lst

def FindOutline(Outlines,pa,pe):
    """
    provide the outline just above the position (of the comment)
    """
    m=None
    for o in Outlines:
        if(o[1]<pa) or ((o[1]==pa)and (o[2]<=pe)):m=o
    return m


if sys.argv[0].upper().find("PYTHON.EXE")>=0:
    del sys.argv[0]

if len(sys.argv)==1:
    print(globals()['__doc__'])
    sys.argv.append(input("pdf file to scan:"))

pdfS=PDF.PdfFileReader(sys.argv[-1])

if '-o' in sys.argv:
    xlFile=sys.argv [sys.argv.index('-o')+1]
else:
    tempFolder=os.environ['TEMP'].replace('\\','/')
    if tempFolder[-1]!='/' : tempFolder+='/'
    xlFile=tempFolder+"Comments on "+os.path.splitext(os.path.split(pdfS.filepath)[-1])[0]+'.xlsx'

#prepare the destination workbook
wb = Workbook()
ws=wb.active
ws.append(('Page','Pos','Chapt','Originator','Comment','Answer'))
ws.column_dimensions[get_column_letter(0+1)].width=5
ws.column_dimensions[get_column_letter(1+1)].width=5
ws.column_dimensions[get_column_letter(2+1)].width=25
ws.column_dimensions[get_column_letter(3+1)].width=15
ws.column_dimensions[get_column_letter(4+1)].width=90
ws.column_dimensions[get_column_letter(5+1)].width=90

#check if decryption is required
if pdfS.isEncrypted: pdfS.decrypt('')

#MyPages will store the matching table page.idnum => pagenumer,page_height
pdfS.MyPages={}

for i,p in enumerate(pdfS.pages):
    pdfS.MyPages[p.indirectRef.idnum]=[i,p['/MediaBox'][3]]

#extract the list of OutLines into MyOutlines
pdfS.MyOutlines=ListOutlines(pdfS)

#extract the comments into MyAnnots
pdfS.MyAnnots=ListAnnots(pdfS)


#sort the comments in the order (Page, vertical position, date)
lst={}
for p in pdfS.MyAnnots.values():
    pp=pdfS.MyPages[p.rawGet("/P").idnum]
    pc=100.0*(1.0-float(int(p['/Rect'][1])/pp[1]))
    lst[(pp[0],pc,p['/M'])]=p

#fill the xl sheet with the comments
for x in sorted(lst.keys()):
    p=lst[x]
    if '/IRT' in p: continue #the comments with IRT are already present in the original comment irt field, we can ignore this one

    #print(x[0],',',end='')
    #print('%.0f %%'%pc,',',end='')
    #print(FindOutline(pdfS.MyOutlines,x[0],x[1])[0],',',end='')
    auth=p['/T']
    if isinstance(auth,bytes):auth=auth.decode('unicode_escape')
    cont=p['/Contents']
    if isinstance(cont,bytes):cont=cont.replace(b'\r',b'\n').decode('unicode_escape')
    #print(cont,',',end='')
    if isinstance(p.irt_str,bytes):p.irt_str=p.irt_str.replace(b'\r',b'\n').decode('unicode_escape')
    #print(p.irt_str)

    ws.append((pdfS.getPageLabel(x[0])  ,'%.0f %%'%pc,FindOutline(pdfS.MyOutlines,x[0],x[1])[0],auth,cont,p.irt_str))

#post insertion formating
for row in ws.iter_rows():
    for cell in row:      
        cell.alignment =  cell.alignment.copy(wrapText=True,vertical='top')

#save and open the file
wb.save(xlFile)
if ('-d' in sys.argv) or ('idlelib.run' in sys.modules):
    os.startfile(xlFile)
